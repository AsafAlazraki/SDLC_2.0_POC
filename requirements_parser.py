"""
requirements_parser.py — parse customer requirement spreadsheets (CSV or Excel)
into a normalised list of requirements ready for the grooming pipeline.

Strategy:
  1. `parse_requirements_file(payload, filename)` — dispatch by extension to
     either CSV or Excel (openpyxl). Returns a raw DataFrame-like list-of-dicts
     plus lightweight metadata (column headers, row count, sheet names).
  2. `autodetect_columns(rows, gemini_api_key)` — one cheap Gemini Flash call
     inspects the first ~5 rows + headers and proposes a mapping from the
     customer's column names onto our canonical fields (id, description,
     priority, source, notes, type). Returns confidence + suggested mapping.
  3. `normalise_rows(rows, mapping)` — apply the mapping, yielding a clean
     `NormalisedRequirement` per row ready for the grooming pipeline.

Never raises for the common cases (bad encoding, empty rows, mystery columns).
Errors come back as `ParseResult.error` with a readable explanation so the UI
can show what went wrong.

Designed for 200–1000 rows per upload. Larger files trigger a
`row_count_warning` in metadata so the caller can decide whether to chunk.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Any, Optional, Tuple

logger = logging.getLogger(__name__)

# Optional deps — soft-imported so the module stays useful in tests without
# Excel support. A missing dep is reported as a clean error at parse time.
try:
    import openpyxl  # type: ignore
except ImportError:
    openpyxl = None  # type: ignore

try:
    from google import genai as _genai  # type: ignore
    from google.genai import types as _genai_types  # type: ignore
except ImportError:
    _genai = None  # type: ignore
    _genai_types = None  # type: ignore


# Canonical fields we map customer columns onto. Grooming agents consume these.
CANONICAL_FIELDS = [
    "id",             # Requirement ID / reference number
    "description",    # The requirement text itself (required)
    "priority",       # High / Med / Low or P1/P2/P3 or similar
    "source",         # Where it came from (stakeholder, doc reference, etc.)
    "type",           # Functional / Non-functional / Bug / etc.
    "notes",          # Free-form additional context
    "acceptance",     # Preliminary acceptance criteria if the customer provided them
    "owner",          # Stakeholder / requestor name
    "tags",           # Comma-separated labels
]

MAX_ROWS_WARN = 1000
MAX_ROWS_HARD = 5000          # absolute cap — beyond this we reject
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB hard cap (typical Excel with 2000 rows ≈ 2–3 MB)


@dataclass
class NormalisedRequirement:
    """One row of the customer's spreadsheet, mapped onto our canonical fields.

    `raw_row` preserves the original dict so downstream consumers can recover
    any column we didn't map (useful for grooming agents that want the full
    customer-verbatim text).
    """
    row_index: int
    id: str = ""
    description: str = ""
    priority: str = ""
    source: str = ""
    type: str = ""
    notes: str = ""
    acceptance: str = ""
    owner: str = ""
    tags: str = ""
    raw_row: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ColumnMapping:
    """Result of LLM column auto-detection."""
    mapping: Dict[str, str] = field(default_factory=dict)   # canonical_field -> source_column
    confidence: str = "low"                                  # "high" | "medium" | "low"
    unmapped_sources: List[str] = field(default_factory=list)
    reasoning: str = ""
    source: str = "autodetect"                               # "autodetect" | "user" | "fallback"


@dataclass
class ParseResult:
    """Container for everything a parse returns.

    Callers always check `error` first; if set, the other fields are best-effort.
    """
    filename: str = ""
    rows: List[Dict[str, Any]] = field(default_factory=list)
    columns: List[str] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)   # Excel only
    sheet_used: str = ""                                   # Excel only
    row_count: int = 0
    warnings: List[str] = field(default_factory=list)
    error: str = ""
    kind: str = ""                                         # "csv" | "excel"

    def as_dict(self) -> Dict[str, Any]:
        return asdict(self)


# ─── Parse dispatch ────────────────────────────────────────────────────────

def parse_requirements_file(payload: bytes, filename: str) -> ParseResult:
    """Public entry point. Dispatch by extension and return a ParseResult.

    Pure-sync. Never raises — failures come back as `ParseResult.error`.
    """
    res = ParseResult(filename=filename or "(unnamed)")

    if not payload:
        res.error = "Empty file payload."
        return res
    if len(payload) > MAX_FILE_BYTES:
        res.error = f"File too large ({len(payload):,} bytes) — max {MAX_FILE_BYTES:,}."
        return res

    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _parse_csv(payload, res, delimiter="\t" if name.endswith(".tsv") else ",")
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_excel(payload, res)
    if name.endswith(".xls"):
        res.error = "Legacy .xls not supported — save as .xlsx and retry."
        return res

    # Fallback: try CSV first, then Excel. Useful when MIME lies about type.
    csv_try = _parse_csv(payload, ParseResult(filename=filename), delimiter=",")
    if not csv_try.error and csv_try.row_count > 0:
        return csv_try
    excel_try = _parse_excel(payload, ParseResult(filename=filename))
    if not excel_try.error and excel_try.row_count > 0:
        return excel_try
    res.error = f"Unsupported file type. Expected .csv, .tsv, .xlsx, or .xlsm."
    return res


def _parse_csv(payload: bytes, res: ParseResult, delimiter: str = ",") -> ParseResult:
    """Parse CSV/TSV payload. Decodes as UTF-8 with BOM handling."""
    try:
        # Handle BOM + mixed Windows/Unix line endings
        text = payload.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)
        res.kind = "csv"
        res.rows = [_strip_row(r) for r in rows if _has_any_content(r)]
        res.columns = list(reader.fieldnames or [])
        res.row_count = len(res.rows)
        _apply_size_warnings(res)
        return res
    except Exception as e:
        res.error = f"CSV parse failed: {e}"
        return res


def _parse_excel(payload: bytes, res: ParseResult) -> ParseResult:
    """Parse Excel payload — uses the first sheet (or 'Requirements' if it exists)."""
    if openpyxl is None:
        res.error = "Excel support requires `openpyxl` — add it to requirements.txt and reinstall."
        return res
    try:
        wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        res.sheet_names = wb.sheetnames
        res.kind = "excel"
        # Prefer a sheet called 'Requirements' / 'Sheet1' / first sheet in that order.
        prefer = next((s for s in ["Requirements", "requirements", "Sheet1"] if s in res.sheet_names), res.sheet_names[0])
        res.sheet_used = prefer
        ws = wb[prefer]

        # Read header row (first non-empty row)
        header: List[str] = []
        data_rows: List[Dict[str, Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not header:
                header = [str(c).strip() if c is not None else "" for c in row]
                # If the header row is entirely empty, skip and try next
                if not any(header):
                    header = []
                    continue
                continue
            # Data row — skip if entirely empty
            vals = [("" if c is None else c) for c in row]
            if not any(str(v).strip() for v in vals):
                continue
            d = {}
            for col_name, val in zip(header, vals):
                if col_name:  # ignore unnamed trailing columns
                    d[col_name] = val
            if d:
                data_rows.append(d)
        res.columns = [h for h in header if h]
        res.rows = [_strip_row(r) for r in data_rows if _has_any_content(r)]
        res.row_count = len(res.rows)
        _apply_size_warnings(res)
        return res
    except Exception as e:
        res.error = f"Excel parse failed: {e}"
        return res


def _strip_row(r: Dict[str, Any]) -> Dict[str, Any]:
    """Trim whitespace on string values and drop empty keys."""
    out = {}
    for k, v in r.items():
        if k is None or str(k).strip() == "":
            continue
        if isinstance(v, str):
            v = v.strip()
        out[str(k).strip()] = v
    return out


def _has_any_content(r: Dict[str, Any]) -> bool:
    """True if the row has any non-empty string/value."""
    return any(str(v).strip() for v in r.values() if v is not None)


def _apply_size_warnings(res: ParseResult) -> None:
    """Annotate warnings for large files. Hard-cap beyond MAX_ROWS_HARD."""
    if res.row_count > MAX_ROWS_HARD:
        res.error = (
            f"Too many rows ({res.row_count:,}). Hard limit {MAX_ROWS_HARD:,}. "
            f"Split the file or drop lower-priority rows and retry."
        )
        res.rows = []
        res.row_count = 0
        return
    if res.row_count > MAX_ROWS_WARN:
        res.warnings.append(
            f"Large requirement set ({res.row_count:,} rows). Grooming will take longer and cost more. "
            f"Consider uploading in chunks by epic/module."
        )
    if res.row_count == 0:
        res.warnings.append("No data rows detected. Check the header row and try again.")


# ─── LLM column auto-detection ─────────────────────────────────────────────

_AUTODETECT_PROMPT = """You are a schema mapper. You will be shown the header row and first few data rows of a customer's requirements spreadsheet. Your job is to map each of the customer's column names onto our canonical fields.

Our canonical fields (use EXACTLY these keys in your mapping):
- "id"          — a requirement ID or reference number (e.g. REQ-001, R.1.2)
- "description" — the main requirement text (REQUIRED — must map to something)
- "priority"    — priority / severity / must-have-vs-nice-to-have
- "source"      — stakeholder name, source document, or origin
- "type"        — functional / non-functional / security / performance / bug / etc.
- "notes"       — additional context, comments, free-form
- "acceptance"  — preliminary acceptance criteria if the customer provided them
- "owner"       — requestor / product owner / stakeholder name
- "tags"        — labels, categories, or comma-separated markers

Rules:
- Only include a key in the mapping if you are confident. Omit keys you're unsure about.
- EVERY canonical field maps to AT MOST ONE source column.
- Source columns that don't match any canonical field stay unmapped — list them in `unmapped_sources`.
- "description" MUST be mapped. If you cannot find an obvious description column, pick the longest-text-looking column.
- Confidence: "high" if most obvious fields are clearly labelled (Requirement, Priority, etc.); "medium" if you had to guess on 1-2; "low" if the schema is highly ambiguous.

Return VALID JSON ONLY (no prose, no code fences), matching this schema EXACTLY:

{
  "mapping": {
    "description": "<source column name>",
    "id": "<source column name>",
    ... only include the ones you mapped
  },
  "confidence": "high" | "medium" | "low",
  "unmapped_sources": ["<source col>", ...],
  "reasoning": "<1-2 sentence explanation>"
}

Customer's header row and first rows follow."""


async def autodetect_columns(
    parse_result: ParseResult,
    gemini_api_key: str,
    *,
    sample_rows: int = 5,
) -> ColumnMapping:
    """Use Gemini Flash to map customer column names onto canonical fields.

    Pure async. Never raises. On any failure (no key, SDK missing, API error,
    bad JSON), returns a ColumnMapping with `source="fallback"` and a rule-based
    best-effort mapping so the upload path still works.
    """
    fallback = _heuristic_mapping(parse_result)

    if not gemini_api_key:
        fallback.reasoning = "No Gemini API key — used heuristic fallback mapping."
        return fallback
    if _genai is None or _genai_types is None:
        fallback.reasoning = "google-genai SDK not installed — used heuristic fallback."
        return fallback
    if not parse_result.columns or parse_result.row_count == 0:
        fallback.reasoning = "No columns / rows to map."
        return fallback

    # Build a compact sample the LLM can reason over.
    sample = parse_result.rows[:sample_rows]
    sample_block = (
        f"Columns: {parse_result.columns}\n\n"
        f"First {len(sample)} row(s):\n"
        + "\n".join(f"  {i+1}. {json.dumps(r, default=str, ensure_ascii=False)[:400]}"
                    for i, r in enumerate(sample))
    )

    try:
        client = _genai.Client(api_key=gemini_api_key)
        response = await client.aio.models.generate_content(
            model="gemini-2.0-flash",
            contents=f"{_AUTODETECT_PROMPT}\n\n{sample_block}",
            config=_genai_types.GenerateContentConfig(temperature=0.1),
        )
        raw = (response.text or "").strip()
        # Strip code fences if the model ignored our "no fences" instruction.
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.IGNORECASE)
        parsed = json.loads(raw)
    except Exception as e:
        logger.warning(f"Column autodetect failed: {e}")
        fallback.reasoning = f"LLM autodetect failed ({type(e).__name__}) — used heuristic fallback."
        return fallback

    # Validate the parsed response against our schema.
    mapping_raw = parsed.get("mapping") or {}
    mapping_clean: Dict[str, str] = {}
    for canonical, source_col in mapping_raw.items():
        if canonical in CANONICAL_FIELDS and isinstance(source_col, str) and source_col in parse_result.columns:
            mapping_clean[canonical] = source_col

    if "description" not in mapping_clean:
        # Fall back to heuristic if the LLM didn't map description
        logger.warning("LLM autodetect skipped description mapping — using heuristic.")
        return fallback

    mapped_sources = set(mapping_clean.values())
    unmapped = [c for c in parse_result.columns if c not in mapped_sources]

    return ColumnMapping(
        mapping=mapping_clean,
        confidence=parsed.get("confidence", "medium") if parsed.get("confidence") in ("high", "medium", "low") else "medium",
        unmapped_sources=unmapped,
        reasoning=(parsed.get("reasoning") or "")[:400],
        source="autodetect",
    )


def _heuristic_mapping(parse_result: ParseResult) -> ColumnMapping:
    """Rule-based fallback — match column names by common substrings."""
    heuristics = {
        "id":          ["id", "ref", "reference", "req", "requirement id", "item"],
        "description": ["description", "requirement", "req", "need", "story", "statement", "detail"],
        "priority":    ["priority", "importance", "severity", "must have", "moscow"],
        "source":      ["source", "origin", "document", "from"],
        "type":        ["type", "category", "kind", "classification"],
        "notes":       ["notes", "comment", "remarks", "context"],
        "acceptance":  ["acceptance", "criteria", "ac"],
        "owner":       ["owner", "requestor", "stakeholder", "po"],
        "tags":        ["tags", "labels", "category"],
    }
    cols_lower = {c.lower(): c for c in parse_result.columns}
    mapping: Dict[str, str] = {}
    for canonical, patterns in heuristics.items():
        for col_lower, col_orig in cols_lower.items():
            # exact match or pattern contained
            if any(col_lower == p or p in col_lower for p in patterns):
                if canonical not in mapping and col_orig not in mapping.values():
                    mapping[canonical] = col_orig
                    break

    # If still no description, pick the column with the longest average string.
    if "description" not in mapping and parse_result.rows:
        best_col = None
        best_avg = 0.0
        for col in parse_result.columns:
            vals = [str(r.get(col, "") or "") for r in parse_result.rows[:20]]
            if not vals:
                continue
            avg = sum(len(v) for v in vals) / len(vals)
            if avg > best_avg:
                best_avg = avg
                best_col = col
        if best_col:
            mapping["description"] = best_col

    used = set(mapping.values())
    return ColumnMapping(
        mapping=mapping,
        confidence="medium" if "description" in mapping else "low",
        unmapped_sources=[c for c in parse_result.columns if c not in used],
        reasoning="Heuristic rule-based mapping (pattern matching on column names).",
        source="fallback",
    )


# ─── Normalisation ──────────────────────────────────────────────────────────

def normalise_rows(
    parse_result: ParseResult,
    mapping: ColumnMapping,
) -> List[NormalisedRequirement]:
    """Apply the mapping to produce NormalisedRequirement objects."""
    inv_mapping = mapping.mapping  # canonical -> source_col
    out: List[NormalisedRequirement] = []
    for i, row in enumerate(parse_result.rows):
        req = NormalisedRequirement(row_index=i, raw_row=dict(row))
        for canonical in CANONICAL_FIELDS:
            source_col = inv_mapping.get(canonical)
            if source_col:
                val = row.get(source_col, "")
                setattr(req, canonical, str(val) if val is not None else "")
        out.append(req)
    return out
